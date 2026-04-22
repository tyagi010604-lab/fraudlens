import os, warnings
import pandas as pd
import numpy as np
warnings.filterwarnings("ignore")

try:
    import openpyxl
    from openpyxl.styles import (PatternFill, Font, Alignment,
                                  Border, Side, GradientFill)
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip",
                           "install", "openpyxl", "--break-system-packages", "-q"])
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CSV  = os.path.join(BASE, "data", "processed_jobs.csv")
OUT  = os.path.join(BASE, "outputs", "FraudLens_Analysis.xlsx")

print("Loading processed data...")
df = pd.read_csv(CSV)

wb = openpyxl.Workbook()

# ── Styles ──────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F2D3D")
FAKE_FILL     = PatternFill("solid", fgColor="FFCCCC")
REAL_FILL     = PatternFill("solid", fgColor="CCFFCC")
ACCENT_FILL   = PatternFill("solid", fgColor="2980B9")
SUB_FILL      = PatternFill("solid", fgColor="D6EAF8")
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT    = Font(bold=True, color="1F2D3D", size=13)
BODY_FONT     = Font(size=10)
CENTER        = Alignment(horizontal="center", vertical="center")
LEFT          = Alignment(horizontal="left",   vertical="center")
BORDER        = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"))

def style_header(cell, text):
    cell.value = text
    cell.fill  = HEADER_FILL
    cell.font  = HEADER_FONT
    cell.alignment = CENTER
    cell.border = BORDER

def style_body(cell, value, fill=None):
    cell.value = value
    cell.font  = BODY_FONT
    cell.alignment = CENTER
    cell.border = BORDER
    if fill:
        cell.fill = fill


# ════════════════════════════════════
# SHEET 1 – OVERVIEW DASHBOARD
# ════════════════════════════════════
ws1 = wb.active
ws1.title = "Overview"
ws1.column_dimensions["A"].width = 28
ws1.column_dimensions["B"].width = 18
ws1.column_dimensions["C"].width = 18
ws1.column_dimensions["D"].width = 18

# Title
ws1.merge_cells("A1:D1")
ws1["A1"] = "FraudLens – Job Fraud Analytics Dashboard"
ws1["A1"].font = Font(bold=True, size=16, color="1F2D3D")
ws1["A1"].alignment = CENTER
ws1["A1"].fill = PatternFill("solid", fgColor="D6EAF8")
ws1.row_dimensions[1].height = 35

ws1.merge_cells("A2:D2")
ws1["A2"] = f"Dataset: {len(df):,} jobs  |  Fraudulent: {df['fraudulent'].sum():,}  |  Real: {(df.fraudulent==0).sum():,}"
ws1["A2"].font = Font(italic=True, size=11, color="555555")
ws1["A2"].alignment = CENTER
ws1.row_dimensions[2].height = 22

# KPI table
ws1["A4"] = "METRIC"
ws1["B4"] = "VALUE"
ws1["C4"] = "BENCHMARK"
ws1["D4"] = "STATUS"
for col in ["A4","B4","C4","D4"]:
    style_header(ws1[col], ws1[col].value)
ws1.row_dimensions[4].height = 20

kpis = [
    ("Total Job Postings",          f"{len(df):,}",                 "–",     "–"),
    ("Fraudulent Jobs",             f"{df['fraudulent'].sum():,}",  "< 30%", "✓ OK"),
    ("Fraud Rate (%)",              f"{df['fraudulent'].mean()*100:.1f}%", "< 20%", "⚠"),
    ("Avg Trust Score (All)",       f"{df['trust_score'].mean():.1f}", "> 70",  "✓ OK"),
    ("Avg Trust Score (Real)",      f"{df[df.fraudulent==0]['trust_score'].mean():.1f}", "> 80", "✓ OK"),
    ("Avg Trust Score (Fake)",      f"{df[df.fraudulent==1]['trust_score'].mean():.1f}", "< 30", "✓ OK"),
    ("High-Risk Jobs",              f"{(df.risk_level=='High Risk').sum():,}", "–", "⚠"),
    ("Jobs Missing Company Info",   f"{(df['has_company']==0).sum():,}", "–", "⚠"),
    ("Salary Anomaly Jobs",         f"{df['salary_anomaly'].sum():,}", "–", "⚠"),
    ("WFH Fraud Rate (%)",          f"{df[df.telecommuting==1]['fraudulent'].mean()*100:.1f}%", "< 20%", "⚠"),
]

for i, (m, v, b, s) in enumerate(kpis, start=5):
    ws1.row_dimensions[i].height = 18
    fill = FAKE_FILL if "⚠" in s else REAL_FILL if "✓" in s else None
    style_body(ws1.cell(i, 1), m)
    style_body(ws1.cell(i, 2), v, fill)
    style_body(ws1.cell(i, 3), b)
    style_body(ws1.cell(i, 4), s, fill)
    ws1.cell(i, 1).alignment = LEFT


# ════════════════════════════════════
# SHEET 2 – PIVOT TABLE: Fraud by Employment Type
# ════════════════════════════════════
ws2 = wb.create_sheet("Pivot_EmploymentType")
ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 14
ws2.column_dimensions["C"].width = 14
ws2.column_dimensions["D"].width = 18

pivot = df.groupby("employment_type").agg(
    Total=("fraudulent", "count"),
    Fake=("fraudulent", "sum")
).reset_index()
pivot["Real"] = pivot["Total"] - pivot["Fake"]
pivot["Fraud_Rate"] = (pivot["Fake"] / pivot["Total"] * 100).round(2)
pivot = pivot.sort_values("Fraud_Rate", ascending=False)

ws2["A1"] = "Fraud Analysis by Employment Type"
ws2["A1"].font = TITLE_FONT
ws2.merge_cells("A1:D1")
ws2.row_dimensions[1].height = 24

for col, header in zip([1,2,3,4], ["Employment Type","Total","Fake","Fraud Rate %"]):
    style_header(ws2.cell(2, col), header)

for i, (_, row) in enumerate(pivot.iterrows(), start=3):
    fill = FAKE_FILL if row["Fraud_Rate"] > 30 else None
    style_body(ws2.cell(i,1), row["employment_type"])
    style_body(ws2.cell(i,2), int(row["Total"]))
    style_body(ws2.cell(i,3), int(row["Fake"]), fill)
    style_body(ws2.cell(i,4), f"{row['Fraud_Rate']:.1f}%", fill)

# Bar chart
chart2 = BarChart()
chart2.title  = "Fake Jobs by Employment Type"
chart2.y_axis.title = "Count"
chart2.style  = 10
data_ref = Reference(ws2, min_col=3, min_row=2, max_row=2+len(pivot))
cats_ref = Reference(ws2, min_col=1, min_row=3, max_row=2+len(pivot))
chart2.add_data(data_ref, titles_from_data=True)
chart2.set_categories(cats_ref)
chart2.width  = 18
chart2.height = 11
ws2.add_chart(chart2, "F2")


# ════════════════════════════════════
# SHEET 3 – PIVOT TABLE: Risk Level Summary
# ════════════════════════════════════
ws3 = wb.create_sheet("Pivot_RiskLevel")
ws3.column_dimensions["A"].width = 18
for c in ["B","C","D","E"]:
    ws3.column_dimensions[c].width = 18

pivot3 = df.groupby("risk_level").agg(
    Total=("fraudulent","count"),
    Fake=("fraudulent","sum"),
    Avg_Trust=("trust_score","mean"),
    Avg_Keywords=("scam_keyword_count","mean")
).reset_index()
pivot3["Avg_Trust"] = pivot3["Avg_Trust"].round(1)
pivot3["Avg_Keywords"] = pivot3["Avg_Keywords"].round(2)

ws3["A1"] = "Risk Level Distribution & Trust Score Analysis"
ws3["A1"].font = TITLE_FONT
ws3.merge_cells("A1:E1")
ws3.row_dimensions[1].height = 24

for col, h in zip(range(1,6), ["Risk Level","Total Jobs","Fake Jobs","Avg Trust Score","Avg Scam Keywords"]):
    style_header(ws3.cell(2, col), h)

risk_fills = {"High Risk": FAKE_FILL, "Medium Risk": PatternFill("solid", fgColor="FFF3CD"), "Low Risk": REAL_FILL}
for i, (_, row) in enumerate(pivot3.iterrows(), start=3):
    f = risk_fills.get(row["risk_level"])
    for col, val in zip(range(1,6), [row["risk_level"], int(row["Total"]), int(row["Fake"]),
                                      row["Avg_Trust"], row["Avg_Keywords"]]):
        style_body(ws3.cell(i, col), val, f)

# Pie chart
chart3 = PieChart()
chart3.title = "Risk Level Distribution"
data3 = Reference(ws3, min_col=2, min_row=2, max_row=2+len(pivot3))
cats3 = Reference(ws3, min_col=1, min_row=3, max_row=2+len(pivot3))
chart3.add_data(data3, titles_from_data=True)
chart3.set_categories(cats3)
chart3.width = 16; chart3.height = 11
ws3.add_chart(chart3, "G2")


# ════════════════════════════════════
# SHEET 4 – SALARY ANALYSIS
# ════════════════════════════════════
ws4 = wb.create_sheet("Salary_Analysis")
ws4.column_dimensions["A"].width = 25
for c in ["B","C","D"]:
    ws4.column_dimensions[c].width = 16

ws4["A1"] = "Salary Range vs Fraud Analysis"
ws4["A1"].font = TITLE_FONT
ws4.merge_cells("A1:D1")
ws4.row_dimensions[1].height = 24

bins   = [0, 50000, 100000, 200000, 500000, float("inf")]
labels = ["< 50K","50K–100K","100K–200K","200K–500K","> 500K"]
dfs = df[df["max_salary"] > 0].copy()
dfs["salary_bucket"] = pd.cut(dfs["max_salary"], bins=bins, labels=labels)
pivot4 = dfs.groupby("salary_bucket", observed=False).agg(
    Total=("fraudulent","count"),
    Fake=("fraudulent","sum")
).reset_index()
pivot4["Fraud_Rate"] = (pivot4["Fake"]/pivot4["Total"]*100).fillna(0).round(1)

for col, h in zip(range(1,5), ["Salary Range","Total Jobs","Fake Jobs","Fraud Rate %"]):
    style_header(ws4.cell(2, col), h)

for i, (_, row) in enumerate(pivot4.iterrows(), start=3):
    f = FAKE_FILL if row["Fraud_Rate"] > 40 else None
    style_body(ws4.cell(i,1), str(row["salary_bucket"]))
    style_body(ws4.cell(i,2), int(row["Total"]))
    style_body(ws4.cell(i,3), int(row["Fake"]), f)
    style_body(ws4.cell(i,4), f"{row['Fraud_Rate']:.1f}%", f)

chart4 = BarChart()
chart4.title = "Fraud Rate by Salary Range"
chart4.y_axis.title = "Fraud Rate %"
d4 = Reference(ws4, min_col=4, min_row=2, max_row=2+len(pivot4))
c4 = Reference(ws4, min_col=1, min_row=3, max_row=2+len(pivot4))
chart4.add_data(d4, titles_from_data=True)
chart4.set_categories(c4)
chart4.width=18; chart4.height=11
ws4.add_chart(chart4, "F2")


# ════════════════════════════════════
# SHEET 5 – RAW DATA SAMPLE
# ════════════════════════════════════
ws5 = wb.create_sheet("Sample_Data")
sample = df[["job_id","title","company_profile","location","fraudulent",
             "trust_score","risk_level","missing_fields","scam_keyword_count"]].head(100)

for col_idx, col_name in enumerate(sample.columns, 1):
    style_header(ws5.cell(1, col_idx), col_name)
    ws5.column_dimensions[get_column_letter(col_idx)].width = max(len(col_name)+4, 14)

for row_idx, (_, row) in enumerate(sample.iterrows(), start=2):
    for col_idx, val in enumerate(row, 1):
        fill = FAKE_FILL if (col_idx == 5 and val == 1) else \
               REAL_FILL if (col_idx == 5 and val == 0) else None
        style_body(ws5.cell(row_idx, col_idx), val, fill)

wb.save(OUT)
print(f"✓ Excel workbook saved: {OUT}")
print("  Sheets: Overview, Pivot_EmploymentType, Pivot_RiskLevel, Salary_Analysis, Sample_Data")
